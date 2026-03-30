"""
reporte_excel.py
================
Responsabilidad: generar el archivo Excel con las tres hojas.

  - Hoja 1: Resumen (KPIs)
  - Hoja 2: Detalle Cambios  (con todos los campos enriquecidos)
  - Hoja 3: Accion Urgente   (solo alertas ALTA)

Grupo IDENTIFICACION en Detalle Cambios:
    Item ID | Fecha Cambio | Cuenta | Proveedor | Marca | SKU GME |
    Ventas  | Fecha Creacion | Estado | Permalink
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT
from procesador import detalle_periodo, ALERTA_STYLE, ORDER_ALERTA


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------


def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def brd():
    t = Side(style="thin", color="E2E8F0")
    return Border(left=t, right=t, top=t, bottom=t)


def ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def lft():
    return Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Hoja 1 – Resumen
# ---------------------------------------------------------------------------


def _hoja_resumen(wb, registros, periodo_det):
    conteo = {"ALTA": 0, "MEDIA": 0, "BAJA": 0, "INFO": 0}
    for r in registros:
        conteo[r["alerta"]] += 1

    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:J2")
    ws["B2"] = "Reporte Cambios de Dimensiones - MercadoLibre"
    ws["B2"].font = Font(name="Arial", bold=True, size=16, color="1E293B")
    ws["B2"].alignment = lft()
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:J3")
    ws["B3"] = f"Periodo analizado:  {periodo_det}"
    ws["B3"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws["B3"].fill = fill("1E2330")
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 10

    kpi_row = 6
    kpis = [
        ("Total cambios", len(registros), "334155"),
        ("ALTA (>=30%)", conteo["ALTA"], "FF4D6D"),
        ("MEDIA (10-29%)", conteo["MEDIA"], "F59E0B"),
        ("BAJA (<10%)", conteo["BAJA"], "22C55E"),
        ("INFO (sin prev.)", conteo["INFO"], "60A5FA"),
    ]
    kpi_cols = ["B", "D", "F", "H", "J"]
    for col, (label, val, bg) in zip(kpi_cols, kpis):
        c_val = ws[f"{col}{kpi_row}"]
        c_val.value = val
        c_val.font = Font(name="Arial", bold=True, size=24, color="FFFFFF")
        c_val.fill = fill(bg)
        c_val.alignment = ctr()
        c_val.border = brd()

        c_lbl = ws[f"{col}{kpi_row + 1}"]
        c_lbl.value = label
        c_lbl.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c_lbl.fill = fill(bg)
        c_lbl.alignment = ctr()
        c_lbl.border = brd()

    ws.row_dimensions[kpi_row].height = 38
    ws.row_dimensions[kpi_row + 1].height = 32

    ley_row = 10
    ws.merge_cells(f"B{ley_row}:J{ley_row}")
    ws[f"B{ley_row}"] = "CRITERIOS DE CLASIFICACION"
    ws[f"B{ley_row}"].font = Font(name="Arial", bold=True, size=9, color="64748B")
    ws.row_dimensions[ley_row].height = 16

    leyendas = [
        (
            "ALTA",
            "Cambio >= 30% en peso facturable -> Reclamar o ajustar urgente",
            "FF4D6D",
        ),
        ("MEDIA", "Cambio 10-29% en peso facturable -> Revisar y decidir", "F59E0B"),
        ("BAJA", "Cambio < 10% en peso facturable -> Monitorear", "22C55E"),
        ("INFO", "Sin datos previos suficientes para comparar", "60A5FA"),
    ]
    for i, (badge, desc, color) in enumerate(leyendas):
        r = ley_row + 1 + i
        ws[f"B{r}"] = badge
        ws[f"B{r}"].font = Font(name="Arial", bold=True, size=9, color=color)
        ws[f"B{r}"].alignment = lft()
        ws.merge_cells(f"C{r}:J{r}")
        ws[f"C{r}"] = desc
        ws[f"C{r}"].font = Font(name="Arial", size=9, color="475569")
        ws[f"C{r}"].alignment = lft()
        ws.row_dimensions[r].height = 14

    note_row = ley_row + 6
    ws.merge_cells(f"B{note_row}:J{note_row}")
    ws[f"B{note_row}"] = (
        "Peso facturable = MAX(peso fisico, peso volumetrico)   |   "
        "Peso volumetrico (kg) = Ancho x Alto x Largo / 5000"
    )
    ws[f"B{note_row}"].font = Font(name="Arial", italic=True, size=9, color="94A3B8")

    ws.column_dimensions["A"].width = 2
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 14


# ---------------------------------------------------------------------------
# Hoja 2 – Detalle Cambios
# ---------------------------------------------------------------------------
#
# Mapa de columnas (1-based):
#   1          → CLASIFICACION  (Alerta)
#   2  – 11   → IDENTIFICACION (Item ID, Fecha Cambio, Cuenta, Proveedor,
#                                Marca, SKU GME, Ventas, Fecha Creacion,
#                                Estado, Permalink)
#   12 – 13   → PESO FISICO (g)
#   14 – 15   → ALTO (cm)
#   16 – 17   → ANCHO (cm)
#   18 – 19   → LARGO (cm)
# ---------------------------------------------------------------------------

# Indices (1-based) de columnas de texto libre → alineación izquierda
_COLS_TEXTO = {4, 5, 6, 7, 10}  # Cuenta, Proveedor, Marca, SKU GME, Estado
# Indices de columnas numéricas enteras (pesos)
_COLS_PESO = {12, 13}


def _hoja_detalle(wb, registros, periodo_det):
    ws = wb.create_sheet("Detalle Cambios")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # --- Fila 1: periodo ---
    total_cols = 19
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"Periodo:  {periodo_det}"
    ws["A1"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws["A1"].fill = fill("1E2330")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 18

    # --- Fila 2: grupos ---
    grupos = [
        (1, 1, "CLASIFICACION", "334155"),
        (2, 11, "IDENTIFICACION", "334155"),
        (12, 13, "PESO FISICO (g)", "1D4ED8"),
        (14, 15, "ALTO (cm)", "065F46"),
        (16, 17, "ANCHO (cm)", "7C3AED"),
        (18, 19, "LARGO (cm)", "9D174D"),
    ]
    ws.row_dimensions[2].height = 18
    for cs, ce, label, bg in grupos:
        col_s = get_column_letter(cs)
        col_e = get_column_letter(ce)
        if cs != ce:
            ws.merge_cells(f"{col_s}2:{col_e}2")
        c = ws[f"{col_s}2"]
        c.value = label
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = ctr()
        c.border = brd()

    # --- Fila 3: encabezados de columna ---
    COLS = [
        # CLASIFICACION
        ("Alerta", 12),
        # IDENTIFICACION
        ("Item ID", 20),
        ("Fecha Cambio", 15),
        ("Cuenta", 22),
        ("Proveedor", 22),
        ("Marca", 18),
        ("SKU GME", 18),
        ("Ventas", 10),
        ("Fecha Creacion", 17),
        ("Estado", 14),
        ("Permalink", 30),
        # PESO FISICO
        ("Peso Ant. (g)", 14),
        ("Peso Nuevo (g)", 14),
        # ALTO
        ("Alto Ant.", 11),
        ("Alto Nvo.", 11),
        # ANCHO
        ("Ancho Ant.", 12),
        ("Ancho Nvo.", 12),
        # LARGO
        ("Largo Ant.", 12),
        ("Largo Nvo.", 12),
    ]
    ws.row_dimensions[3].height = 30
    for ci, (h, w) in enumerate(COLS, start=1):
        cl = get_column_letter(ci)
        c = ws[f"{cl}3"]
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = fill("1E2330")
        c.alignment = ctr()
        c.border = brd()
        ws.column_dimensions[cl].width = w

    # --- Filas de datos ---
    for ri, r in enumerate(registros, start=4):
        ws.row_dimensions[ri].height = 16
        row_bg = "FFFFFF" if ri % 2 == 0 else "F8FAFC"
        a_bg, a_fg = ALERTA_STYLE[r["alerta"]]

        fila = [
            # col 1  – CLASIFICACION
            r["alerta"],
            # cols 2-11 – IDENTIFICACION
            r["item_id"],
            r["fecha"],
            r["cuenta"],
            r["proveedor"],
            r["marca"],
            r["sku_gme"],
            r["ventas"],
            r["fecha_creacion"],
            r["estado"],
            r["permalink"],
            # cols 12-19 – dimensiones
            r["peso_old"],
            r["peso_new"],
            r["alto_old"],
            r["alto_new"],
            r["ancho_old"],
            r["ancho_new"],
            r["largo_old"],
            r["largo_new"],
        ]

        for ci, val in enumerate(fila, start=1):
            cl = get_column_letter(ci)
            c = ws[f"{cl}{ri}"]
            c.value = val
            c.border = brd()
            c.font = Font(name="Arial", size=9, color="1E293B")
            c.fill = fill(row_bg)
            c.alignment = ctr()

            if ci == 1:
                c.fill = fill(a_bg)
                c.font = Font(name="Arial", bold=True, size=9, color=a_fg)

            if ci in _COLS_TEXTO:
                c.alignment = lft()

            if ci in _COLS_PESO and val is not None:
                c.number_format = "#,##0"

    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}{3 + len(registros)}"


# ---------------------------------------------------------------------------
# Hoja 3 – Acción Urgente
# ---------------------------------------------------------------------------


def _hoja_urgente(wb, altas, periodo_det):
    ws = wb.create_sheet("Accion Urgente")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:J1")
    ws["A1"] = (
        "ITEMS QUE REQUIEREN ACCION INMEDIATA -- Cambio >= 30% en peso facturable"
    )
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws["A1"].fill = fill("B91C1C")
    ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Periodo:  {periodo_det}"
    ws["A2"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws["A2"].fill = fill("1E2330")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    COLS3 = [
        ("Item ID", 20),
        ("Fecha Cambio", 15),
        ("Tendencia", 13),
        ("Peso Ant. (g)", 14),
        ("Peso Nvo. (g)", 14),
        ("Fact. Ant. (g)", 14),
        ("Fact. Nvo. (g)", 14),
        ("% Cambio Fact.", 15),
        ("Vol. Ant. (kg)", 13),
        ("Vol. Nvo. (kg)", 13),
    ]
    ws.row_dimensions[3].height = 22
    for ci, (h, w) in enumerate(COLS3, start=1):
        cl = get_column_letter(ci)
        c = ws[f"{cl}3"]
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = fill("1E2330")
        c.alignment = ctr()
        c.border = brd()
        ws.column_dimensions[cl].width = w

    for ri, r in enumerate(altas, start=4):
        ws.row_dimensions[ri].height = 16
        row_bg = "FFFFFF" if ri % 2 == 0 else "FFF5F5"
        es_alza = r["tendencia"] == "alza"
        tend_lbl = "ALZA" if es_alza else "BAJA"
        tend_fg = "B91C1C" if es_alza else "15803D"
        tend_bg = "FEE2E2" if es_alza else "DCFCE7"
        pf = (r["pct_fact"] / 100) if r["pct_fact"] is not None else None

        fila3 = [
            r["item_id"],
            r["fecha"],
            tend_lbl,
            r["peso_old"],
            r["peso_new"],
            r["fact_old"],
            r["fact_new"],
            pf,
            r["vol_old"],
            r["vol_new"],
        ]
        for ci, val in enumerate(fila3, start=1):
            cl = get_column_letter(ci)
            c = ws[f"{cl}{ri}"]
            c.value = val
            c.border = brd()
            c.alignment = ctr()
            c.font = Font(name="Arial", size=9, color="1E293B")
            c.fill = fill(row_bg)

            if ci == 3:
                c.fill = fill(tend_bg)
                c.font = Font(name="Arial", bold=True, size=9, color=tend_fg)
            if ci == 8 and val is not None:
                c.number_format = '+0.0%;-0.0%;"-"'
                c.fill = fill(tend_bg)
                c.font = Font(name="Arial", bold=True, size=9, color=tend_fg)
            if ci in (4, 5, 6, 7) and val is not None:
                c.number_format = "#,##0"
            if ci in (9, 10) and val is not None:
                c.number_format = "0.00"

    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS3))}{3 + len(altas)}"


# ---------------------------------------------------------------------------
# Función pública
# ---------------------------------------------------------------------------


def generar_excel(registros, fecha_inicio, fecha_fin, ruta_salida):
    """
    Genera el Excel completo y devuelve la ruta del archivo guardado.
    """
    os.makedirs(OUTPUT, exist_ok=True)
    wb = Workbook()
    periodo_det = detalle_periodo(fecha_inicio, fecha_fin)
    altas = [r for r in registros if r["alerta"] == "ALTA"]

    _hoja_resumen(wb, registros, periodo_det)
    _hoja_detalle(wb, registros, periodo_det)
    _hoja_urgente(wb, altas, periodo_det)

    sufijo = (
        str(fecha_inicio)
        if fecha_inicio == fecha_fin
        else f"{fecha_inicio}_al_{fecha_fin}"
    )

    nombre_archivo = os.path.join(ruta_salida, f"reporte_dimensiones_{sufijo}.xlsx")
    wb.save(nombre_archivo)
    return nombre_archivo
